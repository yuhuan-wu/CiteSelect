#!/usr/bin/env python3
from __future__ import annotations

import csv
import hashlib
import json
import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from award_matcher import AcademicAwardsMatcher, get_awards_matcher
from llm_client import create_llm_client_from_config
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from utils import (
    ROOT,
    active_publications_jsonl,
    REPORTS_DIR,
    citations_paths_for_paper,
    citations_ref_paths_for_paper,
    legacy_citations_paths,
    legacy_citations_ref_paths,
    load_yaml,
    setup_dirs,
)


TOP_VENUE_KEYWORDS = (
    "ieee transactions on pattern analysis and machine intelligence",
    "ieee transactions on image processing",
    "ieee transactions on neural networks and learning systems",
    "ieee transactions on multimedia",
    "computer vision and pattern recognition",
    "ieee international conference on computer vision",
    "european conference on computer vision",
    "neural information processing systems",
    "international conference on machine learning",
    "international conference on learning representations",
    "aaai conference on artificial intelligence",
    "acm multimedia",
    "international journal of computer vision",
    "ijcv",
    "medical image analysis",
    "ieee transactions on medical imaging",
    "nature communications",
    "science advances",
    "nature machine intelligence",
    "nature medicine",
    "cvpr",
    "iccv",
    "eccv",
    "neurips",
    "nips",
    "icml",
    "iclr",
    "aaai",
    "ijcai",
    "acm mm",
)

HIGH_VENUE_KEYWORDS = (
    "pattern recognition",
    "machine intelligence research",
    "computer vision and image understanding",
    "international joint conference on artificial intelligence",
    "image and vision computing",
    "neurocomputing",
    "expert systems with applications",
    "knowledge-based systems",
    "applied soft computing",
    "engineering applications of artificial intelligence",
    "journal of visual communication and image representation",
    "biomedical signal processing and control",
    "computers in biology and medicine",
    "comput biol medicine",
    "the visual computer",
    "remote sensing",
    "diagnostics",
    "neural networks",
    "neural computing applications",
    "acm trans multim comput commun appl",
    "ieee transactions on circuits and systems for video technology",
    "ieee transactions on instrumentation and measurement",
    "ieee transactions on geoscience and remote sensing",
    "ieee journal of selected topics in applied earth observations and remote sensing",
    "aaai conference on artificial intelligence",
)

MAINSTREAM_VENUE_KEYWORDS = (
    "ieee",
    "acm",
    "springer",
    "elsevier",
    "wiley",
    "aaai",
    "ijcai",
    "nature",
    "science",
    "cell",
)

PREPRINT_VENUE_KEYWORDS = (
    "arxiv",
    "medrxiv",
    "biorxiv",
    "ssrn",
    "research square",
)


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(value)
    except Exception:
        return default


def _normalize_name(name: str) -> str:
    return " ".join((name or "").lower().strip().split())


def _normalize_person_name(name: str) -> str:
    text = (name or "").strip().lower()
    text = text.replace("-", " ")
    text = re.sub(r"[().,]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _normalize_venue_name(venue: str) -> str:
    text = (venue or "").strip().lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[().,:/-]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _load_jsonl(path: Path) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if not path.exists():
        return rows
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            try:
                rec = json.loads(line)
            except Exception:
                continue
            if isinstance(rec, dict):
                rows.append(rec)
    return rows


def _resolve_path(path_str: str) -> Path:
    path = Path(path_str)
    if path.is_absolute():
        return path
    return ROOT / path


def _load_json(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return data if isinstance(data, dict) else {}


def _save_json(path: Path, data: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _load_author_profiles(cfg: Dict[str, Any]) -> Dict[str, List[str]]:
    export_cfg = cfg.get("export") or {}
    path_str = str(export_cfg.get("author_profiles_path") or "").strip()
    if not path_str:
        return {}
    path = _resolve_path(path_str)
    if not path.exists():
        return {}

    records: List[Dict[str, Any]] = []
    if path.suffix.lower() in (".yaml", ".yml"):
        data = load_yaml(path)
        if isinstance(data, dict):
            if isinstance(data.get("authors"), list):
                records = [x for x in data.get("authors") or [] if isinstance(x, dict)]
            else:
                for key, value in data.items():
                    if isinstance(value, dict):
                        rec = dict(value)
                        rec.setdefault("name", key)
                        records.append(rec)
        elif isinstance(data, list):
            records = [x for x in data if isinstance(x, dict)]
    elif path.suffix.lower() == ".json":
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            data = None
        if isinstance(data, dict) and isinstance(data.get("authors"), list):
            records = [x for x in data.get("authors") or [] if isinstance(x, dict)]
        elif isinstance(data, list):
            records = [x for x in data if isinstance(x, dict)]
    elif path.suffix.lower() == ".jsonl":
        records = _load_jsonl(path)
    elif path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if isinstance(row, dict):
                    records.append(dict(row))

    out: Dict[str, List[str]] = {}
    for rec in records:
        tags_raw = rec.get("tags") or rec.get("tag") or rec.get("labels") or []
        if isinstance(tags_raw, str):
            tags = [x.strip() for x in tags_raw.split("|") if x.strip()]
        else:
            tags = [str(x).strip() for x in tags_raw if str(x).strip()]
        names: List[str] = []
        for key in ("name", "author", "primary_name"):
            if rec.get(key):
                names.append(str(rec.get(key)))
        aliases = rec.get("aliases") or rec.get("names") or []
        if isinstance(aliases, str):
            aliases = [x.strip() for x in aliases.split("|") if x.strip()]
        for alias in aliases:
            if alias:
                names.append(str(alias))
        for name in names:
            norm = _normalize_name(name)
            if not norm:
                continue
            out.setdefault(norm, [])
            for tag in tags:
                if tag and tag not in out[norm]:
                    out[norm].append(tag)
    return out


def _extract_author_signals(item: Dict[str, Any]) -> List[Dict[str, Any]]:
    signals = item.get("authorSignals")
    if isinstance(signals, list) and signals:
        out: List[Dict[str, Any]] = []
        for sig in signals:
            if isinstance(sig, dict):
                out.append({
                    "name": sig.get("name") or "",
                    "authorId": sig.get("authorId"),
                    "citationCount": sig.get("citationCount"),
                })
        if out:
            return out

    raw_authors = (((item.get("raw") or {}).get("citingPaper") or {}).get("authors") or [])
    out = []
    for author in raw_authors:
        if not isinstance(author, dict):
            continue
        out.append({
            "name": author.get("name") or "",
            "authorId": author.get("authorId") or author.get("id"),
            "citationCount": None,
        })
    if out:
        return out

    authors = ((item.get("citingPaper") or {}).get("authors") or [])
    for name in authors:
        out.append({"name": str(name), "authorId": None, "citationCount": None})
    return out


def _format_authors(signals: List[Dict[str, Any]]) -> str:
    parts: List[str] = []
    for sig in signals:
        name = str(sig.get("name") or "").strip()
        if not name:
            continue
        cnt = sig.get("citationCount")
        parts.append(f"{name} ({cnt})" if isinstance(cnt, int) else name)
    return "; ".join(parts)


def _parse_author_type_labels(author_type: str) -> List[str]:
    return [part.strip() for part in str(author_type or "").split(";") if part.strip()]


def _label_person_name(label: str) -> Optional[str]:
    match = re.search(r"\(([^()]+)\)\s*$", str(label or "").strip())
    if not match:
        return None
    return match.group(1).strip() or None


def _person_matches_author_list(person_name: str, author_names: List[str]) -> bool:
    person_norm = _normalize_person_name(person_name)
    if not person_norm:
        return False
    person_tokens = person_norm.split()
    for author in author_names:
        author_norm = _normalize_person_name(author)
        if not author_norm:
            continue
        if author_norm == person_norm:
            return True
        author_tokens = author_norm.split()
        if person_tokens == author_tokens:
            return True
        if len(person_tokens) == len(author_tokens):
            compatible = True
            for p_tok, a_tok in zip(person_tokens, author_tokens):
                if p_tok == a_tok:
                    continue
                if len(p_tok) == 1 and a_tok.startswith(p_tok):
                    continue
                if len(a_tok) == 1 and p_tok.startswith(a_tok):
                    continue
                compatible = False
                break
            if compatible:
                return True
    return False


def _needs_llm_author_type_filter(authors: str, author_type: str) -> bool:
    labels = _parse_author_type_labels(author_type)
    if not labels:
        return False
    author_names = [part.strip() for part in str(authors or "").split(";") if part.strip()]
    for label in labels:
        person_name = _label_person_name(label)
        if person_name and not _person_matches_author_list(person_name, author_names):
            return True
    return False


def _build_venue_filter_prompt(batch: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    sys_prompt = (
        "你是学术刊会分级助手。任务：根据 venue 名称，把每个 venue 归类到固定类别。"
        "严格输出 NDJSON，每行一个 JSON 对象，不要输出代码块或额外说明。"
        "\n类别只能取以下之一：顶级期刊/会议、高水平期刊/会议、主流期刊/会议、预印本、其他期刊/会议。"
        "\n规则："
        "\n1) 顶级期刊/会议：CV/AI/医学影像领域公认顶尖 venue，例如 TPAMI、TIP、TNNLS、TMI、Computational Visual Media、IJCV、CVPR、ICCV、ECCV、NeurIPS、ICML、ICLR、Medical Image Analysis、Nature Communications 等，或者比较好的中科院SCI一区期刊。"
        "\n2) 高水平期刊/会议：质量较高但通常不归入最顶尖的一线 venue，比如Pattern Recognition、CVIU、neurocomputing、IJCAI、ICME、BMVC或者大多数CCF-B类期刊/会议，或者中科院SCI二区期刊。"
        "\n3) 主流期刊/会议：常见正式期刊或会议，但难言顶级/高水平，比如ICIP、ICPR等大多数CCF-C类期刊/会议，或者中科院SCI三区期刊。"
        "\n4) 预印本：arXiv、medRxiv、bioRxiv 等。"
        "\n5) 如果不确定，宁可保守分到主流期刊/会议或其他期刊/会议，不要滥判顶级。"
    )
    user_lines = ["INPUT:"]
    for row in batch:
        user_lines.append(json.dumps({"i": row["i"], "venue": row["venue"]}, ensure_ascii=False))
    return [
        {"role": "system", "content": sys_prompt},
        {"role": "user", "content": "\n".join(user_lines)},
    ]


def _build_author_type_filter_prompt(batch: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    sys_prompt = (
        "你是学术作者身份校验助手。任务：根据 authors 列表，过滤 candidate_author_type 中明显不属于这些作者的标签。"
        "严格输出 NDJSON，每行一个 JSON 对象，不要输出代码块或额外说明。"
        "\n输入字段："
        "\n- i: 行号"
        "\n- a: authors，分号分隔的作者列表"
        "\n- t: candidate_author_type，数组，每个元素都是原始标签"
        "\n输出字段："
        '\n- i: 行号'
        '\n- k: 要保留的标签数组，元素必须逐字拷贝自输入 t'
        "\n规则："
        "\n1) 只保留能明确对应到 authors 中某个作者姓名的标签。"
        "\n2) 如果标签中的人名和作者列表只是部分相似、疑似不同人、或缺乏足够证据，则删除。宁可保守删除，不要误保留。"
        "\n3) 允许非常常见的姓名写法差异，例如首字母缩写、句点、连字符。"
        "\n4) 只判断标签与 authors 是否是同一人，不评价奖项真假。"
    )
    user_lines = ["INPUT:"]
    for row in batch:
        user_lines.append(json.dumps({
            "i": row["i"],
            "a": row["a"],
            "t": row["t"],
        }, ensure_ascii=False))
    return [
        {"role": "system", "content": sys_prompt},
        {"role": "user", "content": "\n".join(user_lines)},
    ]


def _apply_llm_author_type_filter(rows: List[Dict[str, Any]], cfg: Dict[str, Any]) -> None:
    awards_cfg = cfg.get("awards") or {}
    llm_filter_cfg = awards_cfg.get("llm_filter") or {}
    if not bool(llm_filter_cfg.get("enabled", False)):
        return

    suspicious_rows: List[Tuple[int, Dict[str, Any]]] = []
    for idx, row in enumerate(rows):
        author_type = str(row.get("author_type") or "").strip()
        if author_type in ("", "未识别"):
            continue
        if _needs_llm_author_type_filter(str(row.get("authors") or ""), author_type):
            suspicious_rows.append((idx, row))
    if not suspicious_rows:
        return

    cache_path = _resolve_path(str(llm_filter_cfg.get("cache_path") or "outputs/cache/award_llm_filter_cache.json"))
    cache = _load_json(cache_path)
    client = create_llm_client_from_config(cfg)
    if not client.api_key:
        logging.warning("LLM author type filter 已启用，但未找到 API key，跳过")
        return

    batch_size = max(1, _safe_int(llm_filter_cfg.get("batch_size"), 40))
    max_output_tokens = max(200, _safe_int(llm_filter_cfg.get("max_output_tokens"), 1200))

    def cache_key(authors: str, author_type: str) -> str:
        return hashlib.sha1(f"{authors}\n{author_type}".encode("utf-8")).hexdigest()

    pending_batches: List[List[Tuple[int, Dict[str, Any], str]]] = []
    current_batch: List[Tuple[int, Dict[str, Any], str]] = []
    for row_idx, row in suspicious_rows:
        key = cache_key(str(row.get("authors") or ""), str(row.get("author_type") or ""))
        cached = cache.get(key)
        if isinstance(cached, list):
            kept = [str(x) for x in cached if str(x).strip()]
            row["author_type"] = "; ".join(kept) if kept else "未识别"
            continue
        current_batch.append((row_idx, row, key))
        if len(current_batch) >= batch_size:
            pending_batches.append(current_batch)
            current_batch = []
    if current_batch:
        pending_batches.append(current_batch)

    for batch in pending_batches:
        prompt_batch = []
        prompt_index_to_row: Dict[int, Tuple[int, Dict[str, Any], str]] = {}
        allowed_labels_by_index: Dict[int, List[str]] = {}
        for prompt_idx, (row_idx, row, key) in enumerate(batch):
            labels = _parse_author_type_labels(str(row.get("author_type") or ""))
            prompt_batch.append({
                "i": prompt_idx,
                "a": str(row.get("authors") or ""),
                "t": labels,
            })
            prompt_index_to_row[prompt_idx] = (row_idx, row, key)
            allowed_labels_by_index[prompt_idx] = labels

        messages = _build_author_type_filter_prompt(prompt_batch)
        content = client.chat(messages, temperature=None, extra_payload={"max_tokens": max_output_tokens})
        if not content:
            logging.warning("LLM author type filter 无返回，保留原始 author_type")
            continue

        parsed: Dict[int, List[str]] = {}
        for line in content.splitlines():
            line = line.strip()
            if not line or line.startswith("```"):
                continue
            try:
                rec = json.loads(line)
            except Exception:
                continue
            if not isinstance(rec, dict):
                continue
            idx = rec.get("i")
            kept = rec.get("k")
            if isinstance(idx, int) and isinstance(kept, list):
                allowed = set(allowed_labels_by_index.get(idx, []))
                parsed[idx] = [str(x).strip() for x in kept if str(x).strip() in allowed]

        for prompt_idx, (_, row, key) in prompt_index_to_row.items():
            kept = parsed.get(prompt_idx)
            if kept is None:
                continue
            row["author_type"] = "; ".join(kept) if kept else "未识别"
            cache[key] = kept

    _save_json(cache_path, cache)


def _venue_type_local(venue: str) -> Tuple[str, bool]:
    text = _normalize_venue_name(venue)
    if not text:
        return "未知", True
    if any(key in text for key in PREPRINT_VENUE_KEYWORDS):
        return "预印本", True
    if any(key in text for key in TOP_VENUE_KEYWORDS):
        return "顶级期刊/会议", True
    if any(key in text for key in HIGH_VENUE_KEYWORDS):
        return "高水平期刊/会议", True
    if any(key in text for key in MAINSTREAM_VENUE_KEYWORDS):
        return "主流期刊/会议", False
    return "其他期刊/会议", False


def _apply_llm_venue_filter(rows: List[Dict[str, Any]], cfg: Dict[str, Any]) -> None:
    export_cfg = cfg.get("export") or {}
    llm_cfg = export_cfg.get("venue_llm_filter") or {}
    if not bool(llm_cfg.get("enabled", False)):
        return

    client = create_llm_client_from_config(cfg)
    if not client.api_key:
        logging.warning("LLM venue filter 已启用，但未找到 API key，跳过")
        return

    cache_path = _resolve_path(str(llm_cfg.get("cache_path") or "outputs/cache/venue_llm_filter_cache.json"))
    cache = _load_json(cache_path)
    batch_size = max(1, _safe_int(llm_cfg.get("batch_size"), 50))
    max_output_tokens = max(200, _safe_int(llm_cfg.get("max_output_tokens"), 1200))

    pending: List[Tuple[int, str, str]] = []
    for idx, row in enumerate(rows):
        venue = str(row.get("citing_venue") or "").strip()
        if not venue:
            continue
        key = hashlib.sha1(venue.encode("utf-8")).hexdigest()
        cached = cache.get(key)
        if isinstance(cached, str) and cached.strip():
            row["venue_type"] = cached.strip()
            continue
        local_label, confident = _venue_type_local(venue)
        row["venue_type"] = local_label
        if not confident:
            pending.append((idx, venue, key))

    if not pending:
        _save_json(cache_path, cache)
        return

    for batch_start in range(0, len(pending), batch_size):
        chunk = pending[batch_start : batch_start + batch_size]
        prompt_batch = [{"i": i, "venue": venue} for i, (_, venue, _) in enumerate(chunk)]
        messages = _build_venue_filter_prompt(prompt_batch)
        content = client.chat(messages, temperature=None, extra_payload={"max_tokens": max_output_tokens})
        if not content:
            continue
        parsed: Dict[int, str] = {}
        for line in content.splitlines():
            line = line.strip()
            if not line or line.startswith("```"):
                continue
            try:
                rec = json.loads(line)
            except Exception:
                continue
            if not isinstance(rec, dict):
                continue
            idx = rec.get("i")
            label = str(rec.get("v") or rec.get("venue_type") or "").strip()
            if isinstance(idx, int) and label in {"顶级期刊/会议", "高水平期刊/会议", "主流期刊/会议", "预印本", "其他期刊/会议"}:
                parsed[idx] = label
        for prompt_idx, (row_idx, venue, key) in enumerate(chunk):
            label = parsed.get(prompt_idx)
            if label:
                rows[row_idx]["venue_type"] = label
                cache[key] = label
            else:
                cache[key] = rows[row_idx].get("venue_type", "其他期刊/会议")

    _save_json(cache_path, cache)


def _author_type(
    signals: List[Dict[str, Any]],
    profiles: Dict[str, List[str]],
    paper_citation_count: int,
    matcher: Optional[AcademicAwardsMatcher] = None,
) -> str:
    manual_tags: List[str] = []
    award_tags: List[str] = []
    max_author_citation = 0
    max_author_name = ""
    for sig in signals:
        name = str(sig.get("name") or "").strip()
        cnt = sig.get("citationCount")
        if isinstance(cnt, int) and cnt > max_author_citation:
            max_author_citation = cnt
            max_author_name = name
        tags = profiles.get(_normalize_name(name), [])
        for tag in tags:
            if tag not in manual_tags:
                manual_tags.append(tag)
        award_matches = None
        if matcher and name:
            award_matches = matcher.match_name(name)
        elif isinstance(sig.get("awardMatches"), list):
            award_matches = sig.get("awardMatches")
        if isinstance(award_matches, list):
            for match in award_matches:
                if not isinstance(match, dict):
                    continue
                award = str(match.get("award") or "").strip()
                matched_name = str(match.get("matched_name") or name).strip()
                if not award:
                    continue
                label = f"{award} ({matched_name})"
                if label not in award_tags:
                    award_tags.append(label)

    auto_tags: List[str] = []
    if max_author_citation >= 10000:
        auto_tags.append(f"高被引学者(S2>={max_author_citation}: {max_author_name})")
    elif max_author_citation >= 3000:
        auto_tags.append(f"较高被引学者(S2>={max_author_citation}: {max_author_name})")
    elif max_author_citation >= 1000:
        auto_tags.append(f"有一定影响力学者(S2>={max_author_citation}: {max_author_name})")

    if paper_citation_count >= 500:
        auto_tags.append(f"高被引引文论文({paper_citation_count})")
    elif paper_citation_count >= 100:
        auto_tags.append(f"较高被引引文论文({paper_citation_count})")

    tags = manual_tags + award_tags + auto_tags
    return "; ".join(tags) if tags else "未识别"


def _citation_role(refined: Dict[str, Any], item: Dict[str, Any]) -> str:
    role = str(refined.get("citation_role") or "").strip()
    if role:
        return role
    text_parts: List[str] = []
    for value in refined.get("key_contexts") or []:
        if isinstance(value, str):
            text_parts.append(value.lower())
    for value in ((item.get("edge") or {}).get("contexts") or []):
        if isinstance(value, str):
            text_parts.append(value.lower())
    text = " ".join(text_parts)
    if any(k in text for k in ("baseline", "compare", "compared", "against", "versus", "outperform")):
        return "性能比较"
    if any(k in text for k in ("based on", "build on", "adopt", "using", "use ", "inspired by")):
        return "方法借鉴"
    if any(k in text for k in ("dataset", "benchmark", "annotated", "annotation", "resource")):
        return "数据/资源引用"
    if text:
        return "背景铺垫"
    return "弱相关"


def _best_context(refined: Dict[str, Any], item: Dict[str, Any]) -> str:
    contexts = refined.get("key_contexts") or ((item.get("edge") or {}).get("contexts") or [])
    parts: List[str] = []
    for value in contexts:
        if isinstance(value, str):
            text = " ".join(value.strip().split())
            if text:
                parts.append(text)
    return " || ".join(parts[:2])


def _worksheet_from_rows(wb: Workbook, title: str, rows: List[Dict[str, Any]], columns: List[Tuple[str, str]]) -> None:
    ws = wb.create_sheet(title=title)
    headers = [header for _, header in columns]
    ws.append(headers)

    fill = PatternFill("solid", fgColor="16324F")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in rows:
        ws.append([row.get(key, "") for key, _ in columns])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    width_map = {
        "A": 34,
        "B": 34,
        "C": 24,
        "D": 34,
        "E": 12,
        "F": 28,
        "G": 18,
        "H": 42,
        "I": 34,
        "J": 18,
        "K": 18,
        "L": 14,
        "M": 12,
        "N": 36,
        "O": 70,
        "P": 10,
        "Q": 24,
        "R": 18,
    }
    for col_letter, width in width_map.items():
        ws.column_dimensions[col_letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _build_row(
    target: Dict[str, Any],
    item: Dict[str, Any],
    refined: Dict[str, Any],
    profiles: Dict[str, List[str]],
    matcher: Optional[AcademicAwardsMatcher] = None,
) -> Dict[str, Any]:
    cp = item.get("citingPaper") or {}
    signals = _extract_author_signals(item)
    citing_paper_citation_count = _safe_int(cp.get("citationCount"))
    importance_score = _safe_int(refined.get("importance_score"))
    row = {
        "target_title": target.get("title") or "",
        "target_venue": target.get("venue") or "",
        "target_year": target.get("year") or "",
        "citing_paper_title": cp.get("title") or refined.get("title") or "",
        "citing_year": cp.get("year") or refined.get("year") or "",
        "citing_venue": cp.get("venue") or refined.get("venue") or "",
        "citing_paper_citation_count": citing_paper_citation_count,
        "authors": _format_authors(signals) or "; ".join(cp.get("authors") or []),
        "author_type": _author_type(signals, profiles, citing_paper_citation_count, matcher=matcher),
        "venue_type": _venue_type_local(str(cp.get("venue") or ""))[0],
        "citation_role": _citation_role(refined, item),
        "importance_score": importance_score,
        "category": refined.get("category") or "",
        "importance_reason": refined.get("importance_reason") or "",
        "citing_content": _best_context(refined, item),
        "is_influential": bool(refined.get("is_influential", (item.get("edge") or {}).get("isInfluential"))),
        "intents": "|".join([str(x) for x in (refined.get("intents") or (item.get("edge") or {}).get("intents") or [])]),
        "url": cp.get("url") or "",
    }
    return row


def export_excel(cfg: Dict[str, Any], only_first_n: Optional[int] = None) -> Path:
    setup_dirs()
    pub_jsonl = active_publications_jsonl()
    if not pub_jsonl.exists():
        raise FileNotFoundError("缺少 publications.jsonl，请先运行 fetch-publications")

    profiles = _load_author_profiles(cfg)
    awards_cfg = cfg.get("awards") or {}
    matcher = get_awards_matcher(cfg) if bool(awards_cfg.get("enabled", True)) else None
    target_papers = _load_jsonl(pub_jsonl)
    if only_first_n:
        target_papers = target_papers[:only_first_n]

    rows: List[Dict[str, Any]] = []
    for target in target_papers:
        paper_id = target.get("paperId")
        if not paper_id:
            continue
        cite_jsonl_new, _ = citations_paths_for_paper(cfg, paper_id)
        cite_jsonl_old, _ = legacy_citations_paths(paper_id)
        cite_jsonl = cite_jsonl_new if cite_jsonl_new.exists() else cite_jsonl_old

        ref_jsonl_new, _ = citations_ref_paths_for_paper(cfg, paper_id)
        ref_jsonl_old, _ = legacy_citations_ref_paths(paper_id)
        ref_jsonl = ref_jsonl_new if ref_jsonl_new.exists() else ref_jsonl_old
        if not cite_jsonl.exists() or not ref_jsonl.exists():
            logging.info(f"跳过 {paper_id}，缺少 citations/refined 文件")
            continue

        raw_items = _load_jsonl(cite_jsonl)
        refined_items = _load_jsonl(ref_jsonl)
        raw_by_cid = {}
        for item in raw_items:
            cid = ((item.get("citingPaper") or {}).get("paperId"))
            if cid:
                raw_by_cid[cid] = item

        for refined in refined_items:
            cid = refined.get("citing_paper_id")
            if not cid:
                continue
            item = raw_by_cid.get(cid)
            if not item:
                continue
            rows.append(_build_row(target, item, refined, profiles, matcher=matcher))

    _apply_llm_author_type_filter(rows, cfg)
    _apply_llm_venue_filter(rows, cfg)

    rows.sort(
        key=lambda row: (
            str(row.get("target_title") or ""),
            -_safe_int(row.get("importance_score")),
            -_safe_int(row.get("citing_paper_citation_count")),
            str(row.get("citing_paper_title") or ""),
        )
    )

    output_cfg = cfg.get("export") or {}
    output_path = _resolve_path(str(output_cfg.get("output_path") or REPORTS_DIR / "representative_citations.xlsx"))
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    columns = [
        ("target_title", "target paper title"),
        ("target_venue", "target venue"),
        ("target_year", "target year"),
        ("citing_paper_title", "citing paper title"),
        ("citing_year", "citing year"),
        ("citing_venue", "citing venue"),
        ("citing_paper_citation_count", "citing paper citations"),
        ("authors", "authors"),
        ("author_type", "author type"),
        ("venue_type", "venue type"),
        ("citation_role", "citation role"),
        ("importance_score", "importance score"),
        ("category", "category"),
        ("importance_reason", "importance reason"),
        ("citing_content", "citing content"),
        ("is_influential", "is influential"),
        ("intents", "intents"),
        ("url", "url"),
    ]
    _worksheet_from_rows(wb, "all_citations", rows, columns)

    top_score_threshold = _safe_int(output_cfg.get("top_score_threshold"), default=4)
    top_rows = [
        row for row in rows
        if _safe_int(row.get("importance_score")) >= top_score_threshold
        or str(row.get("citation_role") or "") in ("方法借鉴", "性能比较", "代表性工作")
    ]
    _worksheet_from_rows(wb, "top_candidates", top_rows, columns)

    wb.save(output_path)
    logging.info(f"Excel 已导出到 {output_path}，共 {len(rows)} 条记录")
    return output_path
